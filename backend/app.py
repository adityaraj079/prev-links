from flask import Flask, jsonify
from openpyxl import load_workbook
from flask_cors import CORS
import requests
from urllib.parse import quote
from google_images_search import GoogleImagesSearch
from bs4 import BeautifulSoup
import re
import base64

app = Flask(__name__)
CORS(app)


# API_KEY = 'c7c44c58ab660dc0237b096e240f5381b99b75efef1205db009c341203bfeb15'

def search_image(query):
    gis = GoogleImagesSearch(API_KEY, CSE_ID)
    gis.search({'q': query})
    if gis.results():
        return gis.results()[0].url
    else:
        return None
    
@app.route('/')  #Decorator (which comes with function)
def welcome():
    return "This is starting page"


def get_image_url_from_site(site_url):
    try:
        # Fetch the HTML content of the website
        response = requests.get(site_url)
        if response.status_code != 200:
            return []

        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all <img> tags on the page
        img_tags = soup.find_all('img')
        # Extract the 'src' attribute of each <img> tag
        image_urls = []

        for img in img_tags:
            src = img.get('src')
            if src and (src.startswith('https://') or src.startswith('http://')):
                if "static" not in src and "logo" not in src and "images" not in src and "icons" not in src and "dvds" not in src and "captcha" not in src:
                    print("Img",len(src))
                    return src
        
        # if img_tags:
        #     image_url = img_tags[0]['src']
        #     return image_url
        # return image_urls
    
    except Exception as e:
        return []
    
def get_video_urls_from_site(site_url):
    try:
        # Fetch the HTML content of the website
        response = requests.get(site_url)
        if response.status_code != 200:
            return []

        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all <video> tags on the page
        video_tags = soup.find_all('video')

        # Extract the 'src' attribute of each <video> tag
        video_urls = re.findall(r'<video[^>]*\sposter=["\']([^"\']*)["\']', response.text)
        print("Vid",len(video_urls))
        return video_urls
    
    except Exception as e:
        return []

def concatenate_characters(characters):
    if isinstance(characters, str):
        # If characters is already a string, return it as is
        return characters
    elif hasattr(characters, '__iter__'):
        # If characters is an iterable (like a list or tuple), join its elements into a single string
        return ''.join(characters)
    else:
        # If characters is neither a string nor an iterable, return an empty string
        return ''


def get_links_with_titles():
    # Load the Excel file
    wb = load_workbook('links-new.xlsx')
    ws = wb['Sheet1']

    # Extract links and titles from Excel
    links_with_titles = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Check if the row has at least 2 columns
        if len(row) >= 2:
            title = row[0]
            link = row[1]
            image_url = row[2] if len(row) > 2 else ''  # Check if image URL is present in the Excel row
            video_url = row[3] if len(row) > 3 else ''  # Check if video URL is present in the Excel row
            
            # Check if image URL and video URL are not already present
            if not image_url:
                # image_url = get_image_url_from_site(link)
                concated_image_url = concatenate_characters(image_url)
                # print('conca ',len(concated_image_url) )
                
                ws.cell(row=idx, column=3, value=concated_image_url)
            
            # If video URL is not present, obtain it from the site
            if not video_url:
                # video_url = get_video_urls_from_site(link)
                concated_video_url = concatenate_characters(video_url)
                ws.cell(row=idx, column=4, value=concated_video_url)

            # Write URLs to the Excel file if they are obtained
            print(title,image_url,video_url)
            # Save changes to the Excel file
            wb.save('links-new.xlsx')

            # Append data to the JSON response
            links_with_titles.append({'title': title, 'link': link, 'image_url': image_url, 'video_url': video_url})

    return jsonify(links_with_titles)

# Define endpoint to fetch links with titles
@app.route('/get_links_with_titles')
def get_links_with_titles_and_images():
    try:
        return get_links_with_titles()
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    

   
@app.route('/get_names', methods=['GET'])
def get_names():
    wb = load_workbook('links-new.xlsx')
    ws = wb['profiles']

    test=0
    names_with_images = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0]:
            # Check if image URL is already present in the Excel file
            if len(row) > 5 and row[5]:  # Check if second column exists and is not empty
                img_url = row[1]
                img_url1 = row[2]
                img_url2 = row[3]
                img_url3 = row[4]
                img_url4 = row[5]
                img_url5 = row[6]
                search_query = quote(row[0])
                follow_link = f'https://www.google.com/search?q={search_query}' + '+porn'
                print(test,f". Image already present for: {row[0]}")
            else:
                # Perform Google search if image URL is not present
                query = row[0] + ' pornstar pornpics portrait'
                query1 = row[0] + ' blacked pichunter thumb'
                query2 = row[0] + ' vixen pornpics portrait'
                query3 = row[0] + ' bangbros pornpics portrait'
                query4 = row[0] + ' innocenthigh pornpics portrait'
                query5 = row[0] + ' naughty america pornpics portrait'
                img_url = search_image(query)
                img_url1 = search_image(query1)
                img_url2 = search_image(query2)
                img_url3 = search_image(query3)
                img_url4 = search_image(query4)
                img_url5 = search_image(query5)
                search_query = quote(row[0])
                follow_link = f'https://www.google.com/search?q={search_query}' + '+porn'
                if img_url:
                    print(test,f". Image found for: {row[0]}")
                    # Write the image URL to the Excel file
                    ws.cell(row=idx, column=2, value=img_url)  # Write image URL in the second column
                    ws.cell(row=idx, column=3, value=img_url1)
                    ws.cell(row=idx, column=4, value=img_url2)
                    ws.cell(row=idx, column=5, value=img_url3)
                    ws.cell(row=idx, column=6, value=img_url4)
                    ws.cell(row=idx, column=7, value=img_url5)
                    wb.save('links-new.xlsx')
                else:
                    print(test,f". No image found for: {row[0]}")
            test+=1
            # Append name and image URL to the list
            names_with_images.append({'id': idx,'name': row[0], 'image': img_url, 'image1': img_url1, 'image2': img_url2, 'image3': img_url3, 'image4': img_url4, 'image5': img_url5, 'follow_link': follow_link})
    return names_with_images

@app.route('/get_video')
def get_video():
    url = "https://www.1024tera.com/sharing/link?surl=Fo5hJjrlmcgyUfBXCM6A7g"
    response = requests.get(url)
    
    # Check if the request was successful
    if response.status_code == 200:
        # Base64 encode the video content
        video_content_base64 = base64.b64encode(response.content).decode('utf-8')
        
        # Create a JSON response with the base64 encoded video content
        response_json = {
            "video_content_base64": video_content_base64
        }
        
        # Return the JSON response
        return jsonify(response_json), 200
    else:
        return jsonify({"error": "Failed to retrieve video", "status_code": response.status_code}), response.status_code

@app.route('/profile/<int:profile_id>')
def get_profile(profile_id):
    try:
        wb = load_workbook('links-new.xlsx')
        ws = wb['profiles']

        profile_data = None
        for row in ws.iter_rows(values_only=True):
            if row[0] and row[0].lower().replace(' ', '-') == f'{profile_id}':
                profile_data = {
                    'id': profile_id,
                    'name': row[1],
                    # 'bio': row[2],
                    # 'image': row[3],
                    # Add other profile details as needed
                }
                break

        if profile_data:
            return jsonify(profile_data), 200
        else:
            return jsonify({'error': 'Profile not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
