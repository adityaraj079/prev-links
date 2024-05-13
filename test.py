from flask import Flask, jsonify
from openpyxl import load_workbook
from flask_cors import CORS
import requests
from urllib.parse import quote

app = Flask(__name__)
CORS(app)

API_KEY = 'c7c44c58ab660dc0237b096e240f5381b99b75efef1205db009c341203bfeb15'
def search_google_image(query):
    params = {
        'q': query,
        'tbm': 'isch',
        'api_key': API_KEY
    }
    response = requests.get('https://serpapi.com/search', params=params)
    data = response.json()
    if 'images_results' in data:
        return data['images_results'][0]['original']  # Return the URL of the first image
    else:
        return None
    
@app.route('/')  #Decorator (which comes with function)
def welcome():
    return "This is starting page"

# Define endpoint to fetch links with titles
@app.route('/get_links_with_titles')
def get_links_with_titles():
    # Load the Excel file
    wb = load_workbook('links-new.xlsx')
    ws = wb['Sheet1']

    # Extract links and titles from Excel
    links_with_titles = []
    for row in ws.iter_rows(values_only=True):
        # Check if the row has the expected number of columns
        if len(row) >= 2:
            links_with_titles.append({'link': row[1], 'title': row[0]})

    return jsonify(links_with_titles)
   
@app.route('/get_names', methods=['GET'])
def get_names():
    wb = load_workbook('links-new.xlsx')
    ws = wb['profiles']
    
    names_with_urls = []
    for row in ws.iter_rows(values_only=True):
        if row[0]:
            # name_parts = row[0].split()
            # name = '+'.join(name_parts)  # Replace space with plus sign
            img_url = search_google_image(row[0])
            
            if img_url:
                name_with_url = {'name': row[0], 'image': img_url}
                names_with_urls.append(name_with_url)
                print(f"Image found for: {row[0]}")
            else:
                print(f"No image found for: {row[0]}")
    
    return jsonify(names_with_urls)



if __name__ == '__main__':
    app.run(debug=True)
